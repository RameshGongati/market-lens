"""Export analysis results to Excel (.xlsx) and PDF (.pdf).

Stage E: rewritten for the two-axis refactor.  The app now produces TWO
result shapes and the exports must handle both:

* **Demand/Supply** — ``all_zones`` (a LIST of zone dicts), ``nearest_demand``
  / ``nearest_supply``, ``trend``, ``status``, ``strength``, ``summary``.
* **Trend Following** — ``strategy == "Trend Following"``, ``signal``,
  ``trend``, ``last_cross`` (dict), ``sma_fast_now`` / ``sma_slow_now``,
  ``status``, ``strength``, ``summary``.

The pre-refactor code assumed a single dict-shaped result and crashed with
``'list' object has no attribute 'items'`` when handed the new shapes (the
dashboard also passed ``list(results.values())`` where a mapping was
expected).  Everything here now accepts a ``dict[str, dict]`` mapping OR a
plain list of result dicts, detects the strategy per stock, and uses
defensive ``.get()`` access throughout — ``.items()`` is never called on
anything that could be a list.
"""

from datetime import datetime
from pathlib import Path
from typing import Any

from utils.logger import get_logger

logger = get_logger(__name__)

# Primary export target: the user's Windows Downloads folder, reachable from
# WSL via the mounted Windows drive — so exported files land somewhere the
# user can open from Windows Explorer.  Change these two constants to retarget.
_WINDOWS_DOWNLOADS = Path("/mnt/c/Users/rames/Downloads")
_WINDOWS_EXPORT_DIR = _WINDOWS_DOWNLOADS / "market-lens"

# Fallback when the Windows mount isn't present (running outside WSL, a
# different machine, or the drive isn't mounted) — keeps exports working.
_FALLBACK_EXPORT_DIR = Path.home() / "market-lens-exports"


def get_export_dir() -> Path:
    """Resolve (and create) the directory export files are written to.

    Resolution order:
      1. ``/mnt/c/Users/rames/Downloads/market-lens`` — but only when the
         Windows Downloads folder actually exists (i.e. the Windows drive is
         mounted under WSL).  The ``market-lens`` subfolder is created if
         needed.
      2. ``~/market-lens-exports`` — the historical location, used as a
         graceful fallback whenever the Windows path is unavailable or can't
         be created.

    Every filesystem touch is wrapped in ``try/except`` so a permissions or
    mount hiccup downgrades to the fallback (logged) instead of crashing the
    export.

    Returns:
        The ``Path`` of the directory to write into — guaranteed to exist on
        return unless even the fallback ``mkdir`` failed (in which case the
        caller's ``save``/``open`` surfaces the real OS error).
    """
    try:
        if _WINDOWS_DOWNLOADS.exists():
            _WINDOWS_EXPORT_DIR.mkdir(parents=True, exist_ok=True)
            return _WINDOWS_EXPORT_DIR
    except Exception as exc:  # noqa: BLE001 — any FS error downgrades to fallback
        logger.warning(
            "Could not use Windows Downloads export dir %s: %s — falling back to %s",
            _WINDOWS_EXPORT_DIR, exc, _FALLBACK_EXPORT_DIR,
        )

    try:
        _FALLBACK_EXPORT_DIR.mkdir(parents=True, exist_ok=True)
    except Exception as exc:  # noqa: BLE001
        logger.error(
            "Failed to create fallback export dir %s: %s",
            _FALLBACK_EXPORT_DIR, exc,
        )
    return _FALLBACK_EXPORT_DIR


# ---------------------------------------------------------------------------
# Result-shape helpers
# ---------------------------------------------------------------------------

def is_trend_following_result(result: dict[str, Any]) -> bool:
    """Return ``True`` when *result* is a Trend Following result dict.

    Detection rule: the Trend Following engine stamps every result with
    ``strategy == "Trend Following"`` (see ``analysis/trend_following.py``).
    Anything else — including legacy results with no ``strategy`` key — is
    treated as Demand/Supply shaped.
    """
    return result.get("strategy") == "Trend Following"


def _normalise_results(
    results: dict[str, dict[str, Any]] | list[dict[str, Any]],
) -> dict[str, dict[str, Any]]:
    """Accept either a ``{symbol: result}`` mapping or a list of result dicts
    and return a uniform ``{symbol: result}`` mapping.

    The dashboard historically passed ``list(results.values())`` — the very
    call that produced ``'list' object has no attribute 'items'``.  Rather
    than relying on every call site being fixed, the exporters normalise
    defensively: lists are keyed by each result's ``symbol`` field (falling
    back to a positional name when absent).
    """
    if isinstance(results, dict):
        return results
    normalised: dict[str, dict[str, Any]] = {}
    for i, r in enumerate(results or []):
        if not isinstance(r, dict):
            continue
        sym = str(r.get("symbol") or f"STOCK_{i + 1}")
        # Strip any exchange suffix (e.g. "RELIANCE.NS" → "RELIANCE") for
        # display keys; keep the full string when there's no dot.
        normalised[sym.split(".")[0] if "." in sym else sym] = r
    return normalised


def _fmt_price(value: Any) -> str:
    """Format a numeric price defensively (``—`` for missing/invalid)."""
    try:
        return f"₹{float(value):,.2f}"
    except (TypeError, ValueError):
        return "—"


def _last_cross_text(result: dict[str, Any]) -> str:
    """Human-readable description of a Trend Following result's last cross."""
    lc = result.get("last_cross") or {}
    cross_type = lc.get("type")
    if not cross_type:
        return "No cross detected"
    candles_ago = lc.get("candles_ago")
    price = lc.get("price")
    parts = [f"{str(cross_type).capitalize()} cross"]
    if candles_ago is not None:
        parts.append(f"{candles_ago} candles ago")
    if price is not None:
        parts.append(f"at {_fmt_price(price)}")
    return " ".join(parts)


# ---------------------------------------------------------------------------
# Excel export
# ---------------------------------------------------------------------------

# Column headers for the per-zone rows on the Details sheet (Demand/Supply).
_ZONE_HEADERS = [
    "Symbol", "Zone Type", "Category", "Proximal", "Distal", "ODD Score",
    "Zone Strength", "Entry Recommendation", "Tradeable", "Trend At Zone",
    "EMA20 Confluence", "Fib Confluence",
]

# Column headers for the per-stock rows on the Details sheet (Trend Following).
_TF_HEADERS = [
    "Symbol", "Signal", "Trend", "Last Cross", "SMA 50", "SMA 200",
    "Status", "Strength", "Summary",
]


def export_to_excel(
    results: dict[str, dict[str, Any]] | list[dict[str, Any]],
    watchlist_name: str,
    analysis_type: str,
    alerts: list[dict[str, Any]] | None = None,
    trading_type: str = "",
    primary_strategy: str = "",
    enhancers: list[str] | None = None,
) -> Path:
    """Export analysis results to an Excel workbook.

    Creates three sheets:

    * **Summary** — one row per stock: symbol, trading type, primary
      strategy, status/signal, strength, price, change %, one-line summary.
    * **Details** — strategy-specific rows: per-zone rows for Demand/Supply
      results, per-stock signal/cross rows for Trend Following results.
    * **Alerts** — alert history (unchanged).

    Both result shapes are handled per stock; ``.items()`` is never called
    on a list (*results* itself may be a list — see ``_normalise_results``).

    Args:
        results: ``{symbol: result}`` mapping or list of result dicts.
        watchlist_name: Name of the watchlist being exported.
        analysis_type: Analysis type / primary strategy that was run
            (kept for backward compatibility with older call sites).
        alerts: Optional list of alert dicts from the database.
        trading_type: The active trading type (e.g. "Short-term Trading").
        primary_strategy: The active primary strategy; falls back to
            *analysis_type* when empty.
        enhancers: The active enhancer selections.

    Returns:
        Path to the generated .xlsx file.

    Raises:
        RuntimeError: If openpyxl is not installed.
    """
    try:
        import openpyxl
        from openpyxl.styles import PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError as exc:
        raise RuntimeError(
            "openpyxl not installed. Run: pip install openpyxl"
        ) from exc

    results_map = _normalise_results(results)
    primary = primary_strategy or analysis_type

    export_dir = get_export_dir()
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = export_dir / f"market_lens_{watchlist_name}_{ts}.xlsx"

    wb = openpyxl.Workbook()

    # --- Sheet 1: Summary ---
    ws_summary = wb.active
    ws_summary.title = "Summary"
    _style_header_row(ws_summary, [
        "Symbol", "Trading Type", "Primary Strategy", "Status/Signal",
        "Strength", "Price (₹)", "Change %", "Summary",
    ])

    status_fills = {
        "bullish": PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid"),
        "bearish": PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid"),
        "neutral": PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid"),
    }

    for symbol, r in results_map.items():
        status = str(r.get("status", "neutral"))
        # Trend Following stocks lead with their BUY/SELL/HOLD signal;
        # D/S stocks show the bullish/bearish/neutral status.
        status_or_signal = (
            str(r.get("signal", "HOLD")) if is_trend_following_result(r)
            else status.capitalize()
        )
        row = [
            symbol,
            trading_type or "—",
            primary or "—",
            status_or_signal,
            str(r.get("strength", "—")),
            r.get("current_price", 0.0),
            r.get("change_pct", 0.0),
            str(r.get("summary", "")),
        ]
        ws_summary.append(row)
        fill = status_fills.get(status)
        if fill:
            for cell in ws_summary[ws_summary.max_row]:
                cell.fill = fill

    _auto_width(ws_summary, get_column_letter)

    # --- Sheet 2: Details (strategy-specific) ---
    ws_detail = wb.create_sheet("Details")
    # A single run uses one strategy for every stock, but detect per stock
    # anyway so a mixed/cached blob can never crash the export.
    any_tf = any(is_trend_following_result(r) for r in results_map.values())
    if any_tf:
        _style_header_row(ws_detail, _TF_HEADERS)
        for symbol, r in results_map.items():
            ws_detail.append([
                symbol,
                str(r.get("signal", "—")),
                str(r.get("trend", "—")),
                _last_cross_text(r),
                r.get("sma_fast_now") if r.get("sma_fast_now") is not None else "—",
                r.get("sma_slow_now") if r.get("sma_slow_now") is not None else "—",
                str(r.get("status", "—")),
                str(r.get("strength", "—")),
                str(r.get("summary", "")),
            ])
    else:
        _style_header_row(ws_detail, _ZONE_HEADERS)
        for symbol, r in results_map.items():
            # all_zones is a LIST of zone dicts — iterate it directly, never
            # call .items() on it (the pre-refactor crash).
            zones = r.get("all_zones") or []
            if not isinstance(zones, list):
                zones = []
            if not zones:
                ws_detail.append([symbol] + ["—"] * (len(_ZONE_HEADERS) - 1))
                continue
            for z in zones:
                if not isinstance(z, dict):
                    continue
                ws_detail.append([
                    symbol,
                    str(z.get("zone_type", "—")),
                    str(z.get("category", "—")),
                    z.get("proximal", "—"),
                    z.get("distal", "—"),
                    z.get("odd_score", "—"),
                    str(z.get("zone_strength", "—")),
                    str(z.get("entry_recommendation", "—")),
                    "Yes" if z.get("is_tradeable") else "No",
                    str(z.get("trend_at_zone", "—")),
                    "Yes" if z.get("ema20_enhancer") else "No",
                    "Yes" if z.get("fib_confluence") else "No",
                ])
    _auto_width(ws_detail, get_column_letter)

    # --- Sheet 3: Alerts ---
    ws_alerts = wb.create_sheet("Alerts")
    _style_header_row(ws_alerts, ["ID", "Stock ID", "Analysis Type", "Message", "Created At"])
    for alert in (alerts or []):
        if not isinstance(alert, dict):
            continue
        ws_alerts.append([
            alert.get("id", ""),
            alert.get("stock_id", ""),
            alert.get("analysis_type", ""),
            alert.get("message", ""),
            str(alert.get("created_at", "")),
        ])
    _auto_width(ws_alerts, get_column_letter)

    wb.save(filename)
    logger.info("Excel exported to %s", filename)
    return filename


# ---------------------------------------------------------------------------
# PDF export
# ---------------------------------------------------------------------------

def export_to_pdf(
    results: dict[str, dict[str, Any]] | list[dict[str, Any]],
    watchlist_name: str,
    analysis_type: str,
    symbol_filter: str | None = None,
    trading_type: str = "",
    primary_strategy: str = "",
    enhancers: list[str] | None = None,
) -> Path:
    """Export analysis results to a PDF report.

    Header shows watchlist, date, trading type, primary strategy, and
    enhancers; then a summary table; then per-stock sections adapted to each
    stock's result shape (zones table for Demand/Supply, signal/cross lines
    for Trend Following).

    Args:
        results: ``{symbol: result}`` mapping or list of result dicts.
        watchlist_name: Name of the watchlist being exported.
        analysis_type: Analysis type / primary strategy that was run
            (kept for backward compatibility with older call sites).
        symbol_filter: If given, export only this single stock.
        trading_type: The active trading type.
        primary_strategy: The active primary strategy; falls back to
            *analysis_type* when empty.
        enhancers: The active enhancer selections.

    Returns:
        Path to the generated .pdf file.

    Raises:
        RuntimeError: If reportlab is not installed.
    """
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import mm
        from reportlab.platypus import (
            SimpleDocTemplate, Table, TableStyle,
            Paragraph, Spacer, HRFlowable,
        )
    except ImportError as exc:
        raise RuntimeError(
            "reportlab not installed. Run: pip install reportlab"
        ) from exc

    results_map = _normalise_results(results)
    primary = primary_strategy or analysis_type

    export_dir = get_export_dir()
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    suffix = f"_{symbol_filter}" if symbol_filter else ""
    filename = export_dir / f"market_lens_{watchlist_name}{suffix}_{ts}.pdf"

    doc = SimpleDocTemplate(str(filename), pagesize=A4)
    styles = getSampleStyleSheet()
    story = []

    # Header — watchlist, date, trading type, primary strategy, enhancers.
    title_style = ParagraphStyle("title", parent=styles["Title"], fontSize=18, spaceAfter=6)
    story.append(Paragraph("Market Lens Analysis Report", title_style))
    enhancer_label = ", ".join(enhancers) if enhancers else "None"
    story.append(Paragraph(
        f"Watchlist: <b>{watchlist_name}</b> &nbsp;|&nbsp; "
        f"Generated: {datetime.now().strftime('%d %b %Y %H:%M')}",
        styles["Normal"],
    ))
    story.append(Paragraph(
        f"Trading Type: <b>{trading_type or '—'}</b> &nbsp;|&nbsp; "
        f"Primary Strategy: <b>{primary or '—'}</b> &nbsp;|&nbsp; "
        f"Enhancers: <b>{enhancer_label}</b>",
        styles["Normal"],
    ))
    story.append(Spacer(1, 6 * mm))
    story.append(HRFlowable(width="100%", thickness=1, color=colors.grey))
    story.append(Spacer(1, 4 * mm))

    filter_results = (
        {symbol_filter: results_map[symbol_filter]}
        if symbol_filter and symbol_filter in results_map
        else results_map
    )

    # Summary table
    if not symbol_filter:
        story.append(Paragraph("Summary", styles["Heading2"]))
        table_data = [["Symbol", "Status/Signal", "Strength", "Price (₹)", "Change %"]]
        for sym, r in filter_results.items():
            status_or_signal = (
                str(r.get("signal", "HOLD")) if is_trend_following_result(r)
                else str(r.get("status", "—")).capitalize()
            )
            try:
                chg_str = f"{float(r.get('change_pct', 0)):+.2f}%"
            except (TypeError, ValueError):
                chg_str = "—"
            table_data.append([
                sym,
                status_or_signal,
                str(r.get("strength", "—")),
                _fmt_price(r.get("current_price")),
                chg_str,
            ])
        t = Table(table_data, hAlign="LEFT")
        t.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#343a40")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f8f9fa")]),
            ("FONTSIZE", (0, 0), (-1, -1), 9),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ]))
        story.append(t)
        story.append(Spacer(1, 6 * mm))

    # Per-stock details — adapted to each stock's result shape.
    for sym, r in filter_results.items():
        status = str(r.get("status", "neutral"))
        strength = str(r.get("strength", "—"))
        color_hex = {"bullish": "#155724", "bearish": "#721c24", "neutral": "#856404"}.get(status, "#000")

        if is_trend_following_result(r):
            signal = str(r.get("signal", "HOLD"))
            story.append(Paragraph(
                f'<font color="{color_hex}"><b>{sym}</b></font> — {signal} | '
                f'Trend: {r.get("trend", "—")} | {strength}',
                styles["Heading3"],
            ))
            story.append(Paragraph(
                f"Price: {_fmt_price(r.get('current_price'))} &nbsp; "
                f"SMA 50: {_fmt_price(r.get('sma_fast_now'))} &nbsp; "
                f"SMA 200: {_fmt_price(r.get('sma_slow_now'))}",
                styles["Normal"],
            ))
            story.append(Paragraph(_last_cross_text(r), styles["Normal"]))
        else:
            story.append(Paragraph(
                f'<font color="{color_hex}"><b>{sym}</b></font> — {status.upper()} | {strength}',
                styles["Heading3"],
            ))
            try:
                chg_str = f"{float(r.get('change_pct', 0)):+.2f}%"
            except (TypeError, ValueError):
                chg_str = "—"
            story.append(Paragraph(
                f"Price: {_fmt_price(r.get('current_price'))} &nbsp; Change: {chg_str}",
                styles["Normal"],
            ))
            # Zone table — all_zones is a LIST of zone dicts.
            zones = r.get("all_zones") or []
            if isinstance(zones, list) and zones:
                zone_rows = [["Type", "Proximal", "Distal", "Score", "Strength", "Tradeable"]]
                for z in zones:
                    if not isinstance(z, dict):
                        continue
                    zone_rows.append([
                        str(z.get("zone_type", "—")),
                        str(z.get("proximal", "—")),
                        str(z.get("distal", "—")),
                        str(z.get("odd_score", "—")),
                        str(z.get("zone_strength", "—")),
                        "Yes" if z.get("is_tradeable") else "No",
                    ])
                zt = Table(zone_rows, hAlign="LEFT")
                zt.setStyle(TableStyle([
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#6c757d")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                    ("FONTSIZE", (0, 0), (-1, -1), 8),
                    ("TOPPADDING", (0, 0), (-1, -1), 3),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
                ]))
                story.append(Spacer(1, 2 * mm))
                story.append(zt)

        recommendation = r.get("recommendation") or r.get("summary", "")
        if recommendation:
            story.append(Spacer(1, 2 * mm))
            for line in str(recommendation).split("\n"):
                if line.strip():
                    story.append(Paragraph(line.strip(), styles["Normal"]))
        story.append(Spacer(1, 4 * mm))
        story.append(HRFlowable(width="100%", thickness=0.5, color=colors.lightgrey))
        story.append(Spacer(1, 3 * mm))

    doc.build(story)
    logger.info("PDF exported to %s", filename)
    return filename


def exports_dir() -> Path:
    """Return the resolved exports directory path (Windows Downloads when
    available, else the home-folder fallback).  Thin alias for
    :func:`get_export_dir` kept for existing call sites (e.g. settings page)."""
    return get_export_dir()


def _style_header_row(ws, headers: list[str]) -> None:
    """Append a bold, dark-background header row to an openpyxl worksheet."""
    from openpyxl.styles import Alignment, Font, PatternFill
    ws.append(headers)
    header_fill = PatternFill(start_color="343A40", end_color="343A40", fill_type="solid")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")


def _auto_width(ws, get_column_letter_fn) -> None:
    """Auto-size column widths based on content."""
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=10)
        ws.column_dimensions[get_column_letter_fn(col[0].column)].width = min(max_len + 4, 60)
