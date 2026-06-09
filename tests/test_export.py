"""Offline tests for the Stage E export rewrite.

Verifies that both exporters handle BOTH result shapes (Demand/Supply with
its LIST of zone dicts, and Trend Following) without raising — in particular
the pre-refactor crash ``'list' object has no attribute 'items'`` must be
gone even when the caller passes a list instead of a mapping.

openpyxl / reportlab generate real files into a temp-redirected exports dir;
no network access is needed anywhere.
"""

from __future__ import annotations

from pathlib import Path
from typing import Any

import pytest

import utils.export as export_mod
from utils.export import (
    _last_cross_text,
    _normalise_results,
    export_to_excel,
    export_to_pdf,
    is_trend_following_result,
)


# ---------------------------------------------------------------------------
# Fixtures — sample results in both shapes
# ---------------------------------------------------------------------------

@pytest.fixture(autouse=True)
def _redirect_exports_dir(tmp_path: Path, monkeypatch: pytest.MonkeyPatch):
    """Write export files into a per-test temp dir, not the user's home."""
    monkeypatch.setattr(export_mod, "_EXPORTS_DIR", tmp_path)
    yield


def _zone(**overrides: Any) -> dict[str, Any]:
    base = {
        "zone_type": "DBR",
        "category": "demand",
        "proximal": 1178.0,
        "distal": 1121.0,
        "odd_score": 6.0,
        "zone_strength": "Strong",
        "entry_recommendation": "Buy near proximal",
        "is_tradeable": True,
        "trend_at_zone": "UP",
        "ema20_enhancer": True,
        "fib_confluence": False,
        "mid": 1149.5,
        "top": 1178.0,
        "bottom": 1121.0,
    }
    base.update(overrides)
    return base


@pytest.fixture
def ds_result() -> dict[str, Any]:
    """Demand/Supply-shaped result — all_zones is a LIST of zone dicts."""
    return {
        "symbol": "RELIANCE.NS",
        "current_price": 1200.50,
        "change_pct": -0.95,
        "change": -11.5,
        "status": "bullish",
        "strength": "Strong",
        "trend": "UP",
        "summary": "Trend: UP | Showing 2 key zones (of 14 detected) | price near fresh demand",
        "all_zones": [_zone(), _zone(zone_type="RBD", category="supply",
                                     proximal=1290.0, distal=1320.0,
                                     is_tradeable=False)],
        "all_zones_count": 14,
        "nearest_demand": _zone(),
        "nearest_supply": None,
        "demand_zones": [_zone()],
        "supply_zones": [],
        "stock_id": 1,
        "exchange": "NSE",
    }


@pytest.fixture
def tf_result() -> dict[str, Any]:
    """Trend Following-shaped result."""
    return {
        "strategy": "Trend Following",
        "symbol": "INFY.NS",
        "current_price": 1500.25,
        "change_pct": 1.2,
        "change": 17.8,
        "trend": "UP",
        "signal": "BUY",
        "last_cross": {"type": "golden", "candles_ago": 8, "price": 1450.0},
        "sma_fast_now": 1480.5,
        "sma_slow_now": 1440.2,
        "status": "bullish",
        "strength": "Strong",
        "summary": "Trend Following | UP | BUY (golden cross 8 candles ago) | 50SMA above 200SMA",
        "stock_id": 2,
        "exchange": "NSE",
    }


# ---------------------------------------------------------------------------
# is_trend_following_result — shape detection
# ---------------------------------------------------------------------------

def test_detect_tf_result(tf_result: dict) -> None:
    assert is_trend_following_result(tf_result) is True


def test_detect_ds_result(ds_result: dict) -> None:
    assert is_trend_following_result(ds_result) is False


def test_detect_empty_result_is_not_tf() -> None:
    assert is_trend_following_result({}) is False


def test_detect_legacy_result_without_strategy_key() -> None:
    """Old cached results have no 'strategy' key — treated as D/S shape."""
    assert is_trend_following_result({"status": "neutral"}) is False


# ---------------------------------------------------------------------------
# _normalise_results — mapping/list duality
# ---------------------------------------------------------------------------

def test_normalise_passes_dict_through(ds_result: dict) -> None:
    m = {"RELIANCE": ds_result}
    assert _normalise_results(m) is m


def test_normalise_converts_list_keyed_by_symbol(ds_result: dict, tf_result: dict) -> None:
    out = _normalise_results([ds_result, tf_result])
    assert set(out) == {"RELIANCE", "INFY"}  # exchange suffix stripped


def test_normalise_handles_missing_symbol() -> None:
    out = _normalise_results([{"status": "neutral"}])
    assert list(out) == ["STOCK_1"]


def test_normalise_empty_list() -> None:
    assert _normalise_results([]) == {}


# ---------------------------------------------------------------------------
# _last_cross_text
# ---------------------------------------------------------------------------

def test_last_cross_text_full(tf_result: dict) -> None:
    text = _last_cross_text(tf_result)
    assert "Golden cross" in text
    assert "8 candles ago" in text


def test_last_cross_text_no_cross() -> None:
    assert _last_cross_text({"last_cross": {"type": None}}) == "No cross detected"


def test_last_cross_text_missing_key() -> None:
    assert _last_cross_text({}) == "No cross detected"


# ---------------------------------------------------------------------------
# Excel export — both shapes, dict and list inputs
# ---------------------------------------------------------------------------

def test_excel_ds_result_dict_input(ds_result: dict) -> None:
    path = export_to_excel(
        {"RELIANCE": ds_result}, "TestWL", "Demand/Supply Zones",
        trading_type="Short-term Trading",
        primary_strategy="Demand/Supply Zones",
        enhancers=["EMA 20 Confluence"],
    )
    assert path.exists() and path.stat().st_size > 0


def test_excel_ds_result_list_input_does_not_raise(ds_result: dict) -> None:
    """THE regression test: a LIST input (the old dashboard call style) with
    a LIST of zones must not raise 'list' object has no attribute 'items'."""
    path = export_to_excel([ds_result], "TestWL", "Demand/Supply Zones")
    assert path.exists()


def test_excel_tf_result(tf_result: dict) -> None:
    path = export_to_excel(
        {"INFY": tf_result}, "TestWL", "Trend Following (SMA50/EMA20)",
        trading_type="Long-term Investment",
        primary_strategy="Trend Following (SMA50/EMA20)",
        enhancers=[],
    )
    assert path.exists() and path.stat().st_size > 0


def test_excel_tf_list_input(tf_result: dict) -> None:
    path = export_to_excel([tf_result], "TestWL", "Trend Following (SMA50/EMA20)")
    assert path.exists()


def test_excel_ds_zone_rows_present(ds_result: dict) -> None:
    """The Details sheet must contain one row per zone."""
    import openpyxl
    path = export_to_excel({"RELIANCE": ds_result}, "TestWL", "Demand/Supply Zones")
    wb = openpyxl.load_workbook(path)
    ws = wb["Details"]
    # Header + 2 zones
    assert ws.max_row == 3
    assert ws.cell(row=2, column=2).value == "DBR"
    assert ws.cell(row=3, column=2).value == "RBD"


def test_excel_tf_details_sheet(tf_result: dict) -> None:
    import openpyxl
    path = export_to_excel({"INFY": tf_result}, "TestWL", "TF")
    wb = openpyxl.load_workbook(path)
    ws = wb["Details"]
    assert ws.max_row == 2  # header + 1 stock row
    assert ws.cell(row=2, column=2).value == "BUY"


def test_excel_ds_result_with_no_zones(ds_result: dict) -> None:
    """A D/S result whose zone list is empty must still export (placeholder row)."""
    ds_result["all_zones"] = []
    path = export_to_excel({"RELIANCE": ds_result}, "TestWL", "D/S")
    assert path.exists()


def test_excel_result_with_error_shape() -> None:
    """An error-only result (insufficient data) must not crash the export."""
    path = export_to_excel(
        {"GHOST": {"error": "Insufficient data.", "status": "neutral"}},
        "TestWL", "D/S",
    )
    assert path.exists()


def test_excel_with_alerts(ds_result: dict) -> None:
    alerts = [{"id": 1, "stock_id": 1, "analysis_type": "D/S",
               "message": "Price near demand zone", "created_at": "2026-06-09"}]
    path = export_to_excel({"RELIANCE": ds_result}, "TestWL", "D/S", alerts=alerts)
    assert path.exists()


# ---------------------------------------------------------------------------
# PDF export — both shapes, dict and list inputs
# ---------------------------------------------------------------------------

def test_pdf_ds_result_dict_input(ds_result: dict) -> None:
    path = export_to_pdf(
        {"RELIANCE": ds_result}, "TestWL", "Demand/Supply Zones",
        trading_type="Short-term Trading",
        primary_strategy="Demand/Supply Zones",
        enhancers=["Fibonacci Confluence", "EMA 20 Confluence"],
    )
    assert path.exists() and path.stat().st_size > 0


def test_pdf_ds_result_list_input_does_not_raise(ds_result: dict) -> None:
    """THE regression test for PDF: list input + list of zones must not raise."""
    path = export_to_pdf([ds_result], "TestWL", "Demand/Supply Zones")
    assert path.exists()


def test_pdf_tf_result(tf_result: dict) -> None:
    path = export_to_pdf(
        {"INFY": tf_result}, "TestWL", "Trend Following (SMA50/EMA20)",
        trading_type="Long-term Investment",
        primary_strategy="Trend Following (SMA50/EMA20)",
    )
    assert path.exists() and path.stat().st_size > 0


def test_pdf_single_stock_filter(ds_result: dict, tf_result: dict) -> None:
    path = export_to_pdf(
        {"RELIANCE": ds_result, "INFY": tf_result},
        "TestWL", "D/S", symbol_filter="RELIANCE",
    )
    assert path.exists()
    assert "RELIANCE" in path.name


def test_pdf_mixed_shapes_do_not_crash(ds_result: dict, tf_result: dict) -> None:
    """A cached blob mixing both shapes must export without raising."""
    path = export_to_pdf(
        {"RELIANCE": ds_result, "INFY": tf_result}, "MixedWL", "D/S",
    )
    assert path.exists()


def test_pdf_result_with_error_shape() -> None:
    path = export_to_pdf(
        {"GHOST": {"error": "Insufficient data.", "status": "neutral"}},
        "TestWL", "D/S",
    )
    assert path.exists()


def test_pdf_nan_price_does_not_crash(tf_result: dict) -> None:
    """Invalid price values must render as '—', never crash the formatter."""
    tf_result["current_price"] = None
    tf_result["sma_fast_now"] = None
    path = export_to_pdf({"INFY": tf_result}, "TestWL", "TF")
    assert path.exists()
